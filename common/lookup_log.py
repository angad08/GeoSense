"""
GeoSense — common/lookup_log.py  (shared)
------------------------------------------
Appends completed lookups to the LookupLogs sheet inside the source workbook.
Shared by v1 and v2 — the log format and rules are identical in both.

Rules (feature spec):
  - Every completed lookup is logged — whether the user gave a Police Station,
    a District, both, or only an address.
  - Both given is NOT a no-op: this is a PS-first system, so the lookup still
    derives the district the PS actually belongs to in the Excel. That derived
    district may NOT be the one the user supplied (and the supplied PS may not
    belong to the supplied district), so the result is always worth logging.
  - One result row   → logged automatically.
  - Many result rows → ask which Sr.No was selected, log only that one.

The sheet is SHARED. These columns are written by the script, located by
header text so the hand-maintained columns interleaved between them are never
disturbed:

  ADDRESS | PREDICTED PS | PREDICTED DISTRICT | RESULT LOOKUP | RESULT MATCH

  PREDICTED PS   — the record the user actually selected, not rank 1. Station
                   name only, so it stays comparable to the hand-entered
                   ACTUAL PS KNOWN and the sheet's MATCH formula keeps working.
  PREDICTED DIST — the district of that record. Station names repeat across
                   districts, so the name alone does not identify it.
  RESULT LOOKUP — what the system figured out:
     Case 1 (PS given)      → "DISTRICT"
     Case 2 (District given) → "POLICE STATION"
     Case 3 (address only)   → "DISTRICT + POLICE STATION"
  RESULT MATCH  — the surety value of the chosen record, in plain words
                  (same wording shown in the output table).
"""

import openpyxl

from common.config import (
    EXCEL_FILE, LOG_SHEET_NAME, LOG_WRITE_COLS,
    LOG_COL_ADDRESS, LOG_COL_PRED_PS, LOG_COL_PRED_DIST,
    LOG_COL_LOOKUP, LOG_COL_MATCH,
)
from common.output import SURETY_LABELS


# Engine case → RESULT LOOKUP wording
LOOKUP_LABELS = {
    1: "DISTRICT",
    2: "POLICE STATION",
    3: "DISTRICT + POLICE STATION",
}


def _select_record(results, interactive):
    """
    One result   → return it directly.
    Many results → ask which Sr.No was put on the website, return that record.

    In non-interactive (one-shot) mode, defaults to the top-ranked result.
    Returns None if the user skips or enters nothing.
    """
    if len(results) == 1:
        return results[0]

    if not interactive:
        return results[0]

    valid = sorted(r["rank"] for r in results)
    while True:
        choice = input(f"  Which Sr.No did you select? {valid} : ").strip()
        if not choice:
            print("  Skipped — row still logged, selection columns left blank.")
            return None
        if choice.isdigit() and int(choice) in valid:
            return next(r for r in results if r["rank"] == int(choice))
        print(f"  Enter one of {valid}, or press Enter to skip.")


def log_lookup(result, address, known_ps, known_district,
               excel_path=EXCEL_FILE, interactive=True):
    """
    Append the chosen record of `result` to the LookupLogs sheet.

    Logged regardless of what the user supplied (PS, District, both, or
    address only), and regardless of whether a record was selected — a
    skipped selection still writes the row, with the record-derived cells
    (PREDICTED PS, PREDICTED DISTRICT, RESULT MATCH) left blank to fill in
    by hand.

    The only thing that stops a write is having nothing to log:
      - No results / unrouted case (case 0) → nothing to log.
    """
    results = result.get("results", [])
    if not results or result.get("case", 0) == 0:
        return

    lookup_label = LOOKUP_LABELS.get(result["case"])
    if lookup_label is None:
        return

    record = _select_record(results, interactive)

    # PREDICTED PS is the record the user actually chose, not rank 1 — with
    # duplicated station names those differ, which is the whole point. The
    # district is written to its own column rather than being concatenated
    # into PREDICTED PS, so PREDICTED PS stays directly comparable to the
    # hand-entered ACTUAL PS KNOWN and the sheet's MATCH formula keeps working.
    #
    # If nothing was selected, the row is still written: the lookup happened
    # and belongs in the log. The cells that depend on a chosen record are
    # left empty for manual completion rather than filled with a guess — the
    # same treatment as the hand-maintained columns.
    match_label = (SURETY_LABELS.get(record.get("confidence", "NONE"), "Unknown")
                   if record else "")

    values = {
        LOG_COL_ADDRESS:   address.strip(),
        LOG_COL_PRED_PS:   record.get("police_station", "") if record else "",
        LOG_COL_PRED_DIST: record.get("district", "") if record else "",
        LOG_COL_LOOKUP:    lookup_label,
        LOG_COL_MATCH:     match_label,
    }

    try:
        _append_row(excel_path, values)
    except Exception as e:
        print(f"  [WARN] Could not write to '{LOG_SHEET_NAME}': {e}")
        return

    if record:
        print(f"\n  [LOG] Saved to '{LOG_SHEET_NAME}': "
              f"{record.get('police_station', '')} | {record.get('district', '')} | "
              f"{lookup_label} | {match_label}")
    else:
        print(f"\n  [LOG] Saved to '{LOG_SHEET_NAME}': {lookup_label} | "
              f"no selection — PREDICTED PS, PREDICTED DISTRICT and RESULT MATCH "
              f"left blank for you to fill")


def _header_columns(ws):
    """Map upper-cased header text -> 1-based column index, from row 1."""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return {
        str(name).strip().upper(): i
        for i, name in enumerate(header, start=1)
        if name is not None and str(name).strip()
    }


def _append_row(excel_path, values):
    """
    Append one row to the log sheet, writing each value into the column whose
    HEADER matches — never by position.

    `values` maps header text -> value. Only those columns are written. The
    manual columns (FILE NO, ACTUAL PS KNOWN, STATUS, MATCH) sit interleaved
    between the script's columns, so a positional ws.append() would shift every
    value left and overwrite hand-entered data. Anything to the right of the
    table — stray unnamed columns, the Accuracy cell — is left untouched.

    A missing expected header raises. Failing loudly is the point: the
    alternative is writing an address into 'FILE NO' and nobody noticing.

    If the sheet carries an Excel Table, its range is extended to include the
    new row. Without that the row lands outside the table, its MATCH formula
    never fills in, and it is silently excluded from the accuracy total.
    """
    wb = openpyxl.load_workbook(excel_path)

    if LOG_SHEET_NAME not in wb.sheetnames:
        raise KeyError(
            f"Sheet '{LOG_SHEET_NAME}' not found in {excel_path}. "
            f"Sheets present: {', '.join(wb.sheetnames)}. "
            f"Nothing was written."
        )

    ws   = wb[LOG_SHEET_NAME]
    cols = _header_columns(ws)

    missing = [c for c in LOG_WRITE_COLS if c.upper() not in cols]
    if missing:
        raise KeyError(
            f"Sheet '{LOG_SHEET_NAME}' is missing expected column(s): "
            f"{', '.join(missing)}. Found: {', '.join(sorted(cols))}. "
            f"Nothing was written — refusing to guess which column to use."
        )

    # First free row, measured across the named columns only so that trailing
    # junk to the right cannot inflate it.
    used     = [cols[c.upper()] for c in LOG_WRITE_COLS] + [
        i for h, i in cols.items() if h not in {c.upper() for c in LOG_WRITE_COLS}
    ]
    last     = 1
    for r in range(2, ws.max_row + 1):
        if any(ws.cell(row=r, column=i).value not in (None, "") for i in used):
            last = r
    target = last + 1

    for header, value in values.items():
        ws.cell(row=target, column=cols[header.upper()], value=value)

    # Keep any Excel Table covering these columns in step with the new row.
    # Note: ws.tables.items() yields (name, ref-string) while ws.tables[name]
    # yields the Table object — mutate the object, never assign a string back,
    # or openpyxl fails on save.
    for name in list(ws.tables):
        tbl = ws.tables[name]
        head, tail  = tbl.ref.split(":")
        col_letters = "".join(ch for ch in tail if ch.isalpha())
        end_row     = int("".join(ch for ch in tail if ch.isdigit()))
        if end_row < target:
            tbl.ref = f"{head}:{col_letters}{target}"

    wb.save(excel_path)
