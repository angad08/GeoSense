"""
GeoSense — lookup_log.py
-------------------------
Appends completed lookups to the LookupResults sheet inside the source workbook.

Rules (feature spec):
  - Every completed lookup is logged — whether the user gave a Police Station,
    a District, both, or only an address.
  - Both given is NOT a no-op: this is a PS-first system, so the lookup still
    derives the district the PS actually belongs to in the Excel. That derived
    district may NOT be the one the user supplied (and the supplied PS may not
    belong to the supplied district), so the result is always worth logging.
  - One result row   → logged automatically.
  - Many result rows → ask which Sr.No was selected, log only that one.

Columns: ADDRESS | RESULT LOOKUP | RESULT MATCH
  RESULT LOOKUP — what the system figured out:
     Case 1 (PS given)      → "DISTRICT"
     Case 2 (District given) → "POLICE STATION"
     Case 3 (address only)   → "DISTRICT + POLICE STATION"
  RESULT MATCH  — the surety value of the chosen record, in plain words
                  (same wording shown in the output table).
"""

import openpyxl

from config import EXCEL_FILE, LOG_SHEET_NAME, LOG_OLD_SHEET, LOG_COLS
from output import SURETY_LABELS


# Engine case → RESULT LOOKUP wording
LOOKUP_LABELS = {
    1: "DISTRICT",
    2: "POLICE STATION",
    3: "DISTRICT + POLICE STATION",
}


def _select_record(results, interactive):
    """
    One result   → return it directly.
    Many results → ask which Sr.No was put on the website, return that record.

    In non-interactive (one-shot) mode, defaults to the top-ranked result.
    Returns None if the user skips or enters nothing.
    """
    if len(results) == 1:
        return results[0]

    if not interactive:
        return results[0]

    valid = sorted(r["rank"] for r in results)
    while True:
        choice = input(f"  Which Sr.No did you select? {valid} : ").strip()
        if not choice:
            print("  Skipped — nothing logged.")
            return None
        if choice.isdigit() and int(choice) in valid:
            return next(r for r in results if r["rank"] == int(choice))
        print(f"  Enter one of {valid}, or press Enter to skip.")


def log_lookup(result, address, known_ps, known_district,
               excel_path=EXCEL_FILE, interactive=True):
    """
    Append the chosen record of `result` to the LookupResults sheet.

    Logged regardless of what the user supplied (PS, District, both, or
    address only). The only thing that stops a write is having nothing to
    log:
      - No results / unrouted case (case 0) → nothing to log.
    """
    results = result.get("results", [])
    if not results or result.get("case", 0) == 0:
        return

    lookup_label = LOOKUP_LABELS.get(result["case"])
    if lookup_label is None:
        return

    record = _select_record(results, interactive)
    if record is None:
        return

    match_label = SURETY_LABELS.get(record.get("confidence", "NONE"), "Unknown")

    try:
        _append_row(excel_path, [address.strip(), lookup_label, match_label])
    except Exception as e:
        print(f"  [WARN] Could not write to '{LOG_SHEET_NAME}': {e}")
        return

    print(f"\n  [LOG] Saved to '{LOG_SHEET_NAME}': "
          f"{lookup_label} | {match_label}")


def _append_row(excel_path, row):
    """
    Open the workbook, ensure the LookupResults sheet exists (renaming the
    legacy 'Sheet1' if present), write a header if the sheet is empty, then
    append `row`. Saves in place; other sheets are preserved untouched.
    """
    wb = openpyxl.load_workbook(excel_path)

    # Rename the legacy sheet on first use; reuse it afterwards.
    if LOG_SHEET_NAME not in wb.sheetnames and LOG_OLD_SHEET in wb.sheetnames:
        wb[LOG_OLD_SHEET].title = LOG_SHEET_NAME

    if LOG_SHEET_NAME in wb.sheetnames:
        ws = wb[LOG_SHEET_NAME]
    else:
        ws = wb.create_sheet(LOG_SHEET_NAME)

    # Add a header only if the sheet has no real first row yet.
    first = next(ws.iter_rows(values_only=True), None)
    if first is None or all(c is None for c in first):
        ws.append(LOG_COLS)

    ws.append(row)
    wb.save(excel_path)
