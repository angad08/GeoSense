"""
GeoSense — geopy_distance.py
-----------------------------
Geocoding and distance calculation.

  geocode_address()     — address string → (lat, lng)
  geodesic_distance()   — (lat, lng) pair → distance in chosen unit
  rank_ps_by_distance() — address + PS list → top N sorted by real distance

geopy.distance.geodesic uses Karney's algorithm on the WGS-84 ellipsoid,
which models Earth's actual (slightly flattened) shape — more accurate than
haversine, which assumes a perfect sphere.

Install dependencies:
    pip install geopy googlemaps
"""

import openpyxl
from geopy.distance import geodesic

from common.api_keys import get_key
from v2.config import (
    TOP_N, EXCEL_FILE, SHEET_NAME, COL_DISTRICT, COL_PS, COL_LAT, COL_LNG,
    GEOCODE_SUFFIX,
)


# ─────────────────────────────────────────────────────────────────────────────
# LAZY CLIENT
# ─────────────────────────────────────────────────────────────────────────────

_gmaps = None


def _client():
    """
    Build the Google Maps client on first use, then reuse it.

    Deferred for the same reason as ai_engine.LazyAgent: the fuzzy and locality
    rungs of the ladder answer many lookups without geocoding anything, and a
    client built at import time would demand GOOGLE_MAPS_API_KEY on runs that
    never make a single Maps call. get_key() raises here — at the first real
    geocode — instead of at import.
    """
    global _gmaps
    if _gmaps is None:
        import googlemaps
        _gmaps = googlemaps.Client(key=get_key("GOOGLE_MAPS_API_KEY"))
    return _gmaps


# ─────────────────────────────────────────────────────────────────────────────
# GEOCODING
# ─────────────────────────────────────────────────────────────────────────────

def geocode_address(address):
    """
    Convert an address string to ((lat, lng), formatted_address).
    Returns None if Google couldn't resolve it.

    Always check the returned formatted_address before trusting the
    coordinates — Google can silently snap to the wrong locality.

    Never raises on API trouble: quota, timeout, transport and API errors are
    caught and reported as None, so the caller falls through the case ladder
    or returns an honest empty result instead of crashing.
    """
    try:
        result = _client().geocode(address)
    except Exception as e:
        print(f"  [WARN] Geocoding API error for '{address}': {e}")
        return None

    if not result:
        return None

    location  = result[0]["geometry"]["location"]
    coords    = (location["lat"], location["lng"])
    formatted = result[0]["formatted_address"]

    return coords, formatted


# ─────────────────────────────────────────────────────────────────────────────
# DISTANCE
# ─────────────────────────────────────────────────────────────────────────────

def geodesic_distance(point1, point2, unit="km"):
    """
    Geodesic (ellipsoidal) distance between two (lat, lng) points.

    Args:
        point1: (latitude, longitude) in decimal degrees
        point2: (latitude, longitude) in decimal degrees
        unit:   "km" | "mi" | "m" | "nmi"

    Returns:
        Distance as float in the requested unit.
    """
    dist = geodesic(point1, point2)

    units = {
        "km":  dist.kilometers,
        "mi":  dist.miles,
        "m":   dist.meters,
        "nmi": dist.nautical,
    }

    if unit not in units:
        raise ValueError(f"Unsupported unit '{unit}'. Choose from {list(units)}.")

    return units[unit]


# ─────────────────────────────────────────────────────────────────────────────
# STATION COORDINATES — STORED IN THE EXCEL ITSELF
# ─────────────────────────────────────────────────────────────────────────────
# Station coordinates are static, so each one is geocoded ONCE and then kept in
# the COL_LAT / COL_LNG columns of the PoliceStation sheet. The Excel is the
# single source of truth — there is no side-car cache file to rebuild or keep
# in sync.
#
# A blank LAT/LNG simply means "not geocoded yet". _fill_missing_coords() fills
# those rows on demand the first time they are needed and writes them back, so
# adding a station to the Excel needs no separate build step. A station that
# fails to geocode is left blank and retried on the next lookup — the failure
# is reported, never silently cached as a permanent dead end.
#
# On a normal lookup, every station already has coordinates, so the only live
# API call is the user's input address.

_coords_cache = None


def _header_index(ws):
    """Map upper-cased header name -> 1-based column index for the first row."""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return {
        str(name).strip().upper(): i
        for i, name in enumerate(header, start=1)
        if name is not None
    }


def _load_coords_cache():
    """
    Lazy-load station coordinates from the PoliceStation sheet into a dict:
        (DISTRICT, PS_NAME) -> (lat, lng)   both values present and numeric
        (DISTRICT, PS_NAME) -> None         blank or unparseable — needs geocoding

    An unreadable workbook yields an empty dict (with a warning) rather than
    raising, so a lookup degrades to "no distance ranking" instead of crashing.
    """
    global _coords_cache
    if _coords_cache is None:
        _coords_cache = {}
        try:
            # Default data_only=False: this workbook holds formulas on other
            # sheets, and we must never round-trip them into static values.
            ws = openpyxl.load_workbook(EXCEL_FILE)[SHEET_NAME]
        except Exception as e:
            print(f"  [WARN] Could not read station coordinates from {EXCEL_FILE}: {e}")
            return _coords_cache

        cols    = _header_index(ws)
        i_dist  = cols.get(COL_DISTRICT.upper())
        i_ps    = cols.get(COL_PS.upper())
        i_lat   = cols.get(COL_LAT.upper())
        i_lng   = cols.get(COL_LNG.upper())

        if not i_dist or not i_ps:
            print(f"  [WARN] '{SHEET_NAME}' is missing the "
                  f"{COL_DISTRICT}/{COL_PS} columns — no coordinates loaded.")
            return _coords_cache

        for row in ws.iter_rows(min_row=2, values_only=True):
            district, ps = row[i_dist - 1], row[i_ps - 1]
            if district is None or ps is None:
                continue

            key = (str(district).strip().upper(), str(ps).strip().upper())

            # LAT/LNG columns may not exist yet on a first run — that is simply
            # "nothing geocoded so far", not an error.
            lat = row[i_lat - 1] if i_lat and i_lat <= len(row) else None
            lng = row[i_lng - 1] if i_lng and i_lng <= len(row) else None

            try:
                _coords_cache[key] = (float(lat), float(lng))
            except (TypeError, ValueError):
                _coords_cache[key] = None      # blank or junk → needs geocoding

    return _coords_cache


def _write_coords(updates):
    """
    Write freshly geocoded coordinates back into the PoliceStation sheet, in a
    single save. `updates` is a list of (district, ps_name, lat, lng).

    The LAT/LNG columns are created on first use. Written with openpyxl so the
    workbook's other sheets — including their formulas and tables — survive
    untouched, the same way lookup_log.py appends to LookupLogs.

    A failed save (most often: the file is open in Excel) is reported and
    swallowed. The coordinates still apply to the current lookup from memory;
    they are simply re-geocoded next time instead of costing a crash.
    """
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME]

        cols   = _header_index(ws)
        i_dist = cols.get(COL_DISTRICT.upper())
        i_ps   = cols.get(COL_PS.upper())
        i_lat  = cols.get(COL_LAT.upper())
        i_lng  = cols.get(COL_LNG.upper())

        # Create the coordinate columns the first time we ever write.
        if not i_lat:
            i_lat = ws.max_column + 1
            ws.cell(row=1, column=i_lat, value=COL_LAT)
        if not i_lng:
            i_lng = ws.max_column + 1
            ws.cell(row=1, column=i_lng, value=COL_LNG)

        # (district, ps) -> sheet row number
        row_of = {}
        for r in range(2, ws.max_row + 1):
            district = ws.cell(row=r, column=i_dist).value
            ps       = ws.cell(row=r, column=i_ps).value
            if district is None or ps is None:
                continue
            row_of.setdefault(
                (str(district).strip().upper(), str(ps).strip().upper()), r
            )

        written = 0
        for district, ps, lat, lng in updates:
            r = row_of.get((str(district).strip().upper(), str(ps).strip().upper()))
            if r is None:
                continue
            ws.cell(row=r, column=i_lat, value=lat)
            ws.cell(row=r, column=i_lng, value=lng)
            written += 1

        wb.save(EXCEL_FILE)
        print(f"  [SAVED] {written} station coordinate(s) written to "
              f"'{SHEET_NAME}' in {EXCEL_FILE.name}")

    except Exception as e:
        print(f"  [WARN] Could not save coordinates to {EXCEL_FILE.name}: {e}")
        print(f"         (Is the file open in Excel? Results are still correct; "
              f"these stations will be geocoded again next time.)")


def _fill_missing_coords(district, ps_names, cache):
    """
    Geocode the stations that have no coordinates yet, update `cache` in place,
    and persist everything that resolved in one save.

    Stations that fail to geocode are left blank in the sheet and reported, so
    they are retried on the next lookup — a transient API problem never becomes
    a permanent hole. A station that fails repeatedly is a sign its name needs
    fixing in the Excel; the warning names it every time so it stays visible.
    """
    print(f"  [INFO] {len(ps_names)} station(s) in {district} have no "
          f"coordinates yet — geocoding once and saving to the Excel.")

    updates = []
    for ps in ps_names:
        query  = f"{ps} Police Station, {district}, {GEOCODE_SUFFIX}"
        result = geocode_address(query)          # returns None on any failure

        if result:
            (lat, lng), _ = result
            cache[(str(district).strip().upper(), str(ps).strip().upper())] = (lat, lng)
            updates.append((district, ps, lat, lng))
        else:
            print(f"  [WARN] Could not geocode station: {query}")

    if updates:
        _write_coords(updates)


class RankedResults(list):
    """
    A plain list of ranked result dicts, plus `excluded`: the station names
    that could not be distance-ranked because they have no coordinates and
    could not be geocoded. Callers read it defensively via
    getattr(ranked, "excluded", []), so anything that substitutes a plain
    list (e.g. the test harness mock) still works unchanged.
    """
    def __init__(self, iterable=(), excluded=None):
        super().__init__(iterable)
        self.excluded = list(excluded or [])


# ─────────────────────────────────────────────────────────────────────────────
# PS RANKING BY REAL DISTANCE
# ─────────────────────────────────────────────────────────────────────────────

def rank_ps_by_distance(input_address, ps_list, district, top_n=TOP_N):
    """
    Geocode the input address, read each Police Station's coordinates from the
    PoliceStation sheet, compute real geodesic distance, and return top_n
    sorted closest first.

    Once every station in the district has coordinates, the input address is
    the only live API call. Stations still missing coordinates are geocoded
    once here and saved back to the Excel, so that cost is paid a single time.

    This replaces the AI distance-estimation that was in ai_engine.py.
    Distance is now a real number from real coordinates, not a guess.

    Args:
        input_address : raw address string from the user
        ps_list       : list of Police Station names (strings) for this district
        district      : district name — cache lookup key alongside each PS name
        top_n         : number of results to return (default: config.TOP_N)

    Returns:
        RankedResults (a list) of dicts:
            { police_station, district, distance, distance_km, resolved_address }
        `distance` is the display string ("~12.3 km"); `distance_km` is the
        unrounded float, kept for sorting and for callers that need to compare
        against a threshold (see engine.py's Case 2 sanity check).
        `.excluded` lists stations skipped because they have no coordinates and
        could not be geocoded — excluded, noted, never guessed.
        Empty if the input address couldn't be geocoded.

    Note on jurisdiction vs distance:
        Nearest by straight-line distance is NOT the same as correct jurisdiction.
        Jurisdiction boundaries are administrative — this ranking is a strong
        signal, not a guaranteed answer. Always return top_n, not just 1.
    """
    cache        = _load_coords_cache()
    district_key = str(district).strip().upper()

    # Stations with no coordinates yet are geocoded once, here, and written back
    # to the Excel — so a station newly added to the sheet just works, with no
    # rebuild step to remember.
    missing = [
        ps for ps in ps_list
        if cache.get((district_key, str(ps).strip().upper())) is None
    ]
    if missing:
        _fill_missing_coords(district, missing, cache)

    # Anything still without coordinates could not be geocoded at all. It is
    # excluded from the ranking and reported by name — never guessed at.
    excluded = [
        ps for ps in ps_list
        if cache.get((district_key, str(ps).strip().upper())) is None
    ]
    if excluded:
        print(f"  [WARN] {len(excluded)} station(s) in {district} could not be "
              f"geocoded — excluded from distance ranking.")

    origin_result = geocode_address(input_address)     # the 1 live API call

    if not origin_result:
        print(f"  [WARN] Could not geocode input address: '{input_address}'")
        return RankedResults([], excluded=excluded)

    origin_coords, _ = origin_result
    results = []

    for ps in ps_list:
        coords = cache.get((district_key, str(ps).strip().upper()))
        if coords is None:
            continue                                   # already in `excluded`

        dist_km = geodesic_distance(origin_coords, coords, unit="km")

        results.append({
            "police_station":   ps,
            "district":         district,
            "distance":         f"~{round(dist_km, 1)} km",
            "distance_km":      dist_km,      # unrounded — for sorting and callers
            "resolved_address": "",           # coords come from the sheet, not a live geocode
        })

    # Sort by real distance, take top N
    results.sort(key=lambda x: x["distance_km"])

    return RankedResults(results[:top_n], excluded=excluded)
